import sys
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = json.loads(sys.argv[1])

provider_name = data.get('providerName', '')
date_deliver  = data.get('dateDeliver', '')
schools       = data.get('schools', [])    
products      = data.get('products', [])   

wb = Workbook()
ws = wb.active
ws.title = 'Espelho'

bold        = Font(bold=True)
bold_white  = Font(bold=True, color='FFFFFF')
fill_header = PatternFill('solid', fgColor='1F4E79')   
fill_total  = PatternFill('solid', fgColor='2E75B6') 
fill_school = PatternFill('solid', fgColor='D6E4F0')   
center      = Alignment(horizontal='center', vertical='center')
thin        = Side(style='thin')
border      = Border(left=thin, right=thin, top=thin, bottom=thin)

ws['A1'] = 'Nome da Cidade'
ws['A2'] = 'Node do departamento'
ws['A3'] = f'FORNECEDOR: {provider_name}'
ws['A4'] = f'DATA DE ENTREGA: {date_deliver}'

for row in [3, 4]:
    ws.cell(row=row, column=1).font = bold

START_ROW = 6
START_COL = 2   

ws.cell(row=START_ROW, column=1, value='ESCOLA').font    = bold_white
ws.cell(row=START_ROW, column=1).fill                    = fill_header
ws.cell(row=START_ROW, column=1).alignment               = center
ws.cell(row=START_ROW, column=1).border                  = border

for col_idx, product in enumerate(products, start=START_COL):
    cell = ws.cell(row=START_ROW, column=col_idx, value=product['name'])
    cell.font      = bold_white
    cell.fill      = fill_header
    cell.alignment = center
    cell.border    = border

total_col = START_COL + len(products)
cell = ws.cell(row=START_ROW, column=total_col, value='TOTAL')
cell.font      = bold_white
cell.fill      = fill_header
cell.alignment = center
cell.border    = border

for row_idx, school in enumerate(schools, start=START_ROW + 1):
    name_cell = ws.cell(row=row_idx, column=1, value=school['name'])
    name_cell.font   = bold
    name_cell.border = border

    row_total = 0
    for col_idx, product in enumerate(products, start=START_COL):
        qty = product['schools'].get(str(school['id']), 0) or 0
        cell = ws.cell(row=row_idx, column=col_idx, value=qty)
        cell.alignment = center
        cell.border    = border
        row_total += qty

    total_cell = ws.cell(row=row_idx, column=total_col, value=row_total)
    total_cell.alignment = center
    total_cell.border    = border

total_row = START_ROW + 1 + len(schools)

cell = ws.cell(row=total_row, column=1, value='TOTAL')
cell.font      = bold_white
cell.fill      = fill_total
cell.alignment = center
cell.border    = border

grand_total = 0
for col_idx, product in enumerate(products, start=START_COL):
    col_total = sum(
        product['schools'].get(str(school['id']), 0) or 0
        for school in schools
    )
    cell = ws.cell(row=total_row, column=col_idx, value=col_total)
    cell.font      = bold_white
    cell.fill      = fill_total
    cell.alignment = center
    cell.border    = border
    grand_total += col_total

cell = ws.cell(row=total_row, column=total_col, value=grand_total)
cell.font      = bold_white
cell.fill      = fill_total
cell.alignment = center
cell.border    = border

ws.column_dimensions['A'].width = 35
for col_idx in range(START_COL, total_col + 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 18

import pathlib
downloads = pathlib.Path.home() / 'Downloads'
output_path = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    'espelho_temp.xlsx'
)
wb.save(output_path)
print(output_path)
